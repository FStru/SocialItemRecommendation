# !/usr/bin/env python
# -*-coding:utf-8 -*-

import openpyxl

def append_xlsx_detail(descrip, result_dict1, result_dict2, result_dict3):
    data = openpyxl.load_workbook('../temp/predict_data_temp.xlsx')
    sheetnames = data.sheetnames
    table = data[sheetnames[0]]
    table = data.active
    nrows = table.max_row
    ncolumns = table.max_column
    nrows = nrows+1
    table.cell(nrows, 1).value = descrip
    nrows = nrows+1
    metrics_keys = list(result_dict1.keys())
    for i in range(len(metrics_keys)):
        table.cell(nrows, 3 * i + 1).value = metrics_keys[i]+"@10"
        table.cell(nrows, 3 * i + 2).value = metrics_keys[i]+"@20"
        table.cell(nrows, 3 * i + 3).value = metrics_keys[i]+"@30"
    nrows = nrows+1
    c = 0
    for key in metrics_keys:
        for value in result_dict1[key]:
            c = c+1
            table.cell(nrows, c).value = value

    nrows = nrows + 1
    c = 0
    for key in metrics_keys:
        for value in result_dict2[key]:
            c = c + 1
            table.cell(nrows, c).value = value

    nrows = nrows + 1
    c = 0
    for key in metrics_keys:
        for value in result_dict3[key]:
            c = c + 1
            table.cell(nrows, c).value = value

    data.save('../temp/predict_data_temp.xlsx')